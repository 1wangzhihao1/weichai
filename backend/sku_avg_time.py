import os
import sys
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook

current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.abspath(os.path.join(current_dir, ".."))
if project_root not in sys.path:
    sys.path.insert(0, project_root)

from backend.database import PartMaster, SessionLocal
from scenarios.order_picking.data_paths import SKU_TIME_DIR


def _first_existing_column(df: pd.DataFrame, names: List[str], fallback_index: int = -1):
    for name in names:
        if name in df.columns:
            return name
    raise ValueError(f"Missing required column by name. Tried names={names}")


def _find_excel_files() -> List[str]:
    files = []
    if SKU_TIME_DIR.is_dir():
        for name in os.listdir(SKU_TIME_DIR):
            lower_name = name.lower()
            if name.startswith("~$") or not lower_name.endswith((".xlsx", ".xls")):
                continue
            files.append(str(SKU_TIME_DIR / name))

    unique_files = []
    seen = set()
    for path in files:
        abs_path = os.path.abspath(path)
        if abs_path not in seen and os.path.exists(abs_path):
            unique_files.append(abs_path)
            seen.add(abs_path)
    return unique_files


def _normalize_sku(value) -> str:
    return str(value or "").strip()


def _load_clean_rows(path: str) -> pd.DataFrame:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        wb.close()
        return pd.DataFrame()

    index_by_name = {str(name).strip(): idx for idx, name in enumerate(header) if name is not None}

    def idx(names: List[str], fallback: int) -> int:
        for name in names:
            if name in index_by_name:
                return index_by_name[name]
        raise ValueError(f"{os.path.basename(path)} missing required column by name: {names}")

    start_idx = idx(["开始时间"], 2)
    end_idx = idx(["结束时间"], 3)
    status_idx = idx(["状态"], 8)
    sku_idx = idx(["SKU"], 9)
    name_idx = idx(["物料名称"], 10)
    picked_qty_idx = idx(["已拣选数量"], 12)
    confirm_idx = idx(["确认代码"], 16)

    records = []
    max_idx = max(start_idx, end_idx, status_idx, sku_idx, name_idx, picked_qty_idx, confirm_idx)
    for row in rows:
        if len(row) <= max_idx:
            continue
        sku = _normalize_sku(row[sku_idx])
        if not sku or sku.lower() == "nan":
            continue
        status = str(row[status_idx] or "").strip()
        confirm = str(row[confirm_idx] or "").strip()
        if status not in {"完成", "确定", ""}:
            continue
        if confirm not in {"确定", "完成", ""}:
            continue
        try:
            qty = float(row[picked_qty_idx] or 0.0)
        except Exception:
            continue
        if qty <= 0:
            continue

        start_time = pd.to_datetime(row[start_idx], errors="coerce")
        end_time = pd.to_datetime(row[end_idx], errors="coerce")
        if pd.isna(start_time) or pd.isna(end_time):
            continue
        duration = (end_time - start_time).total_seconds()
        if duration <= 0:
            continue

        records.append(
            {
                "source_file": os.path.basename(path),
                "SKU": sku,
                "物料名称": str(row[name_idx] or ""),
                "已拣选数量": qty,
                "耗时秒": duration,
                "单件耗时": duration / qty,
            }
        )

    wb.close()
    return pd.DataFrame.from_records(records)


def filter_long_time_outliers(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    q1 = float(df["单件耗时"].quantile(0.25))
    q3 = float(df["单件耗时"].quantile(0.75))
    iqr = q3 - q1
    iqr_upper = q3 + 3.0 * iqr if iqr > 0 else q3
    hard_upper = 300.0
    upper = max(iqr_upper, hard_upper)

    clean = df[df["单件耗时"] <= upper].copy()
    removed = len(df) - len(clean)
    clean.attrs["long_outlier_upper"] = upper
    clean.attrs["long_outlier_removed"] = removed
    clean.attrs["long_outlier_iqr_upper"] = iqr_upper
    clean.attrs["long_outlier_hard_upper"] = hard_upper
    return clean


def build_sku_average_times(files: Optional[List[str]] = None) -> pd.DataFrame:
    source_files = files or _find_excel_files()
    if not source_files:
        raise FileNotFoundError("No picking Excel files found under raw_data/sku_time")

    frames = []
    print("=" * 80)
    print("SKU average picking time rebuild")
    print("=" * 80)
    for path in source_files:
        clean = _load_clean_rows(path)
        frames.append(clean)
        fast_rows = int((clean["单件耗时"] <= 0.5).sum())
        print(f"Source: {path}")
        print(f"  valid rows: {len(clean)}")
        print(f"  unique SKU : {clean['SKU'].nunique()}")
        print(f"  fast rows  : {fast_rows} (unit time <= 0.5s, kept)")

    df_before_long_filter = pd.concat(frames, ignore_index=True)
    df_clean = filter_long_time_outliers(df_before_long_filter)
    upper = float(df_clean.attrs.get("long_outlier_upper", 0.0))
    removed = int(df_clean.attrs.get("long_outlier_removed", 0))
    iqr_upper = float(df_clean.attrs.get("long_outlier_iqr_upper", 0.0))
    hard_upper = float(df_clean.attrs.get("long_outlier_hard_upper", 0.0))
    print("-" * 80)
    print(f"Total valid rows before long-outlier filter: {len(df_before_long_filter)}")
    print(f"Long outlier upper bound: {upper:.3f}s per unit (max of IQR upper {iqr_upper:.3f}s and hard upper {hard_upper:.3f}s)")
    print(f"Removed long-outlier rows: {removed}")
    print(f"Total valid rows after long-outlier filter : {len(df_clean)}")
    print(f"Total unique SKU : {df_clean['SKU'].nunique()}")
    print("Only long-time outliers are filtered; fast full-package picking records are kept.")

    sku_stats = (
        df_clean.groupby("SKU")
        .agg(
            物料名称=("物料名称", "first"),
            来源文件数=("source_file", "nunique"),
            出现行数=("耗时秒", "count"),
            总已拣选数量=("已拣选数量", "sum"),
            总耗时秒=("耗时秒", "sum"),
        )
        .reset_index()
    )
    sku_stats["单件平均耗时秒"] = (sku_stats["总耗时秒"] / sku_stats["总已拣选数量"]).round(3)
    sku_stats = sku_stats.sort_values("单件平均耗时秒", ascending=False).reset_index(drop=True)
    return sku_stats


def sync_to_part_master(sku_stats: pd.DataFrame, dry_run: bool = False) -> Dict[str, int]:
    if dry_run:
        return {"deleted": 0, "inserted": int(len(sku_stats))}

    db = SessionLocal()
    if hasattr(db, "__next__"):
        db = next(db)
    try:
        deleted_rows = db.query(PartMaster).delete()
        inserted = 0
        for _, row in sku_stats.iterrows():
            db.add(
                PartMaster(
                    part_type=str(row["SKU"]).strip(),
                    standard_p_time=float(row["单件平均耗时秒"]),
                )
            )
            inserted += 1
        db.commit()
        return {"deleted": int(deleted_rows), "inserted": inserted}
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def rebuild_sku_avg_time(files: Optional[List[str]] = None, dry_run: bool = False) -> Dict[str, object]:
    """Rebuild SKU standard picking time rows and return a backend-friendly summary."""
    source_files = [os.path.abspath(path) for path in files] if files else _find_excel_files()
    sku_stats = build_sku_average_times(source_files)
    result = sync_to_part_master(sku_stats, dry_run=dry_run)

    return {
        "source_files": source_files,
        "sku_count": int(len(sku_stats)),
        "deleted": int(result["deleted"]),
        "inserted": int(result["inserted"]),
        "dry_run": bool(dry_run),
    }


def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="Rebuild t_part_master from historical picking Excel files, keeping full-package fast picks."
    )
    parser.add_argument(
        "--excel",
        action="append",
        default=None,
        help="Picking Excel path. Can be provided multiple times. Default: all Excel files under raw_data/sku_time.",
    )
    parser.add_argument("--dry-run", action="store_true", help="Compute only; do not update database.")
    args = parser.parse_args()

    files = [os.path.abspath(path) for path in args.excel] if args.excel else None
    sku_stats = build_sku_average_times(files)

    print("-" * 80)
    print("Top 10 slowest SKU:")
    print(sku_stats[["SKU", "物料名称", "单件平均耗时秒", "来源文件数", "出现行数"]].head(10).to_string(index=False))

    d00_count = int(sku_stats["SKU"].astype(str).str.endswith(":D00").sum())
    a01_count = int(sku_stats["SKU"].astype(str).str.endswith(":A01").sum())
    print("-" * 80)
    print(f"SKU suffix count: D00={d00_count}, A01={a01_count}")
    print("Full SKU codes are kept separate, for example 610800050009:D00 != 610800050009:A01.")

    result = sync_to_part_master(sku_stats, dry_run=args.dry_run)
    print("-" * 80)
    if args.dry_run:
        print(f"Dry run complete. Would insert {result['inserted']} SKU rows into t_part_master.")
    else:
        print(f"Database updated. Deleted {result['deleted']} old rows, inserted {result['inserted']} SKU rows.")
    print("=" * 80)


if __name__ == "__main__":
    main()
